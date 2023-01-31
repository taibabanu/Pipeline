import os
from typing import List
import xml.etree.ElementTree as ET
from datetime import datetime
import openpyxl

class TestCaseDetailsParser:
    """
    Class to extract test case details from XML file and write in excel file(.xlsx).

    """

    def __init__(self, file_name: str) -> None:
        """
        Initialization for the TestCaseDetailsParser class

        Parameters
        ----------
        file_name: str
            xml file to parse

        """
        self.file_name = file_name
        self.file_path = os.path.join(
            os.getcwd().rsplit("\\", 1)[0], f"inputs\{file_name}"
        )

    def parse_xml(self) -> List:
        """
        Method to parse xml file to extract test case details.


        Returns
        -------
        `List`
            Data list consists of test case information.

        """
        tree = ET.parse(self.file_path)
        root = tree.getroot()
        data_list = [["Name", "Info Type", "Header Name","Header value"]]
        for tc_spec in root.findall('test_case_specification'):
            test_case_details = tc_spec.find('test_case_specification_details').attrib
            for each_type in ['header', 'uut']:
                tag = tc_spec.find(each_type)
                for each_tr in tag.findall('tr'):
                    temp_list = list()
                    temp_list.append('')
                    temp_list.append(each_type)
                    for each_td in each_tr.findall("td"):
                        value = each_td.find('br').text
                        if value and value.startswith('tcs'):
                            tc_name = value
                        temp_list.append(value)
                    temp_list[0] = tc_name
                    data_list.append(temp_list)
            for key in test_case_details:
                data_list.append([tc_name, "test_case_details", key, test_case_details[key]])
        self._write_to_excel(data_list)

    def _write_to_excel(self, data_list: List) -> None:
        """
        Method to write the test case details in the excel file(.xlsx).

        Parameters
        ----------
        `data_list`
            Data list consists of test case information.

        """
        wb = openpyxl.Workbook()
        sheet = wb.active
        row = 1
        for each_data_list in data_list:
            sheet.cell(row=row, column=1, value=each_data_list[0])
            sheet.cell(row=row, column=2, value=each_data_list[1])
            sheet.cell(row=row, column=3, value=each_data_list[2])
            sheet.cell(row=row, column=4, value=each_data_list[3])
            row += 1
        current_time = datetime.now().strftime("%Y-%m-%d-%H%M%S")
        wb.save(
            os.path.join(os.getcwd().rsplit("\\", 1)[0], f"outputs\extracted_xml{current_time}.xls")
        )

